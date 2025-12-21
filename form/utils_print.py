import openpyxl
from openpyxl.styles import Alignment, Font
import win32com.client as win32
import os
from datetime import datetime, timedelta
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

tipo=""
def gerar_ac(dados, caminho_pdf_original="TemplateAC.xlsx"):
    def adicionar_dia_util(data):
        data += timedelta(days=1)
        if data.weekday() == 5:  # sábado
            data += timedelta(days=2)
        elif data.weekday() == 6:  # domingo
            data += timedelta(days=1)
        return data

    caminho_template = "TemplateAC.xlsx"
    wb = openpyxl.load_workbook(caminho_template)
    ws = wb["Template Formulário"]

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = "portrait"
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3

    
    ws["F7"] = dados.get("tag")
    ws["F8"] = dados.get("certificado")
    ws["C9"] = dados.get("data")
    ws["C8"] = dados.get("sistema")

    #Tratamento do campo Local
    local = (dados.get("local") or "").strip()

    if len(local) > 28:
        i = local.find(" ", 28)
        if i != -1:
            local = local[:i] + "\n" + local[i+1:]

    ws["C7"] = local
    ws["C7"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["C"].width = 22
    ws.row_dimensions[7].height = 15 * max(1, local.count("\n") + 1)

    
    tag = (dados.get("tag") or "").upper()
   
    if (
        tag.startswith("TE") or
        tag.endswith("-TE") or
        "-TE-" in tag
    ):
        ws["B2"] = "Análise Crítica de Calibração dos Sensores de Temperatura"
        tipo = "TE"

    elif (
        tag.startswith(("TT", "TIT", "TI")) or
        tag.endswith(("-TT", "-TIT", "-TI")) or
        any(x in tag for x in ("-TT-", "-TIT-", "-TI-"))
    ):
        ws["B2"] = "Análise Crítica de Calibração dos Sensores de Temperatura"
        tipo = "TT"

    elif (
        tag.startswith(("PT", "PIT")) or
        tag.endswith(("-PT", "-PIT")) or
        any(x in tag for x in ("-PT-", "-PIT-"))
    ):
        ws["B2"] = "Análise Crítica de Calibração dos Transmissores de Pressão"
        tipo = "PT"

    else:
        ws["B2"] = "Análise Crítica de Calibração dos Transmissores de Pressão Diferencial"
        tipo = "DPT"


    #Report Date (+1 dia) - útil
    if dados.get("report_date"):
        dt = datetime.strptime(dados["report_date"], "%d/%m/%Y")
        dt_util = adicionar_dia_util(dt)
        ws["H40"] = dt_util.strftime("%d/%m/%Y")
    
    
    range_atualizado = dados.get("range_atualizado", False)
    ns_atualizado = dados.get("sn_atualizado", False)

    blocos = []

    if range_atualizado:
        blocos.append(
            TextBlock(
                text="( X ) Sim (  ) Não\nOBSERVAÇÕES:\n",
                font=InlineFont()
            )
        )
        blocos.append(
            TextBlock(
                text="Novo range e alarmes alterados no computador de vazão\n",
                font=InlineFont(b=True)
            )
        )

    elif ns_atualizado:
        blocos.append(
            TextBlock(
                text="( X ) Sim (  ) Não\nOBSERVAÇÕES:\n",
                font=InlineFont()
            )
        )
        blocos.append(
            TextBlock(
                text="Novo NS alterado no computador de vazão / XML / SFP\n",
                font=InlineFont(b=True)
            )
        )
        

    else:
        blocos.append(
            TextBlock(
                text="(  ) Sim ( X ) Não\nOBSERVAÇÕES:\n",
                font=InlineFont()
            )
        )

    # montar rich text final
    rich = CellRichText(*blocos)

    # escrever na célula B35
    ws["B35"].value = rich
    ws["B35"].alignment = Alignment(wrap_text=True, vertical="top")

   
    wb.save(caminho_template)

    # Definir o caminho de saída do PDF
    pasta_saida = os.path.dirname(os.path.abspath(caminho_pdf_original))
    certificado = dados.get("certificado", "").replace(" ", "")
    tag_limpa = dados.get("tag", "").replace(" ", "")
    nome_pdf = f"{certificado}_{tag_limpa}_AC.pdf"
    caminho_pdf_final = os.path.join(pasta_saida, nome_pdf)
    if os.path.exists(caminho_pdf_final):
        try:
            os.remove(caminho_pdf_final)
        except PermissionError:
            raise PermissionError(
                f"⚠ O arquivo PDF está aberto e não pode ser sobrescrito:\n{caminho_pdf_final}"
            )

  
    excel = win32.DispatchEx("Excel.Application") 
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.Interactive = False

    try:
        wb_excel = excel.Workbooks.Open(os.path.abspath(caminho_template))
        wb_excel.ExportAsFixedFormat(0, caminho_pdf_final)
        wb_excel.Close(SaveChanges=False)

    finally:
        excel.Quit()  

    return caminho_pdf_final,tipo